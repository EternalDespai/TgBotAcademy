import io
import re
from collections import defaultdict, Counter
import openpyxl

THEME_REGEX = re.compile(r"^Урок\s*№\s*\d+\.\s*Тема:\s*.+$", re.IGNORECASE)

def detect_excel_type(data: bytes) -> str:
    if len(data) >= 2 and data[0:2] == b"PK": return "xlsx"
    return "unknown"

# --- Метод 1: Расписание ---
def report_schedule_count(wb) -> str:
    counter = Counter()
    ws = wb.worksheets[0]
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str) and "Предмет:" in cell:
                for line in cell.splitlines():
                    line = line.strip()
                    if line.startswith("Предмет:"):
                        subj = line.replace("Предмет:", "", 1).strip()
                        if subj: counter[subj] += 1

    if not counter: return "Не нашел строк 'Предмет:'."

    lines = ["📊 <b>Количество пар по предметам:</b>\n"]
    for name, cnt in counter.most_common():
        lines.append(f"▫️ {name}: <b>{cnt}</b>")
    return "\n".join(lines)


# --- Метод 2: Темы уроков ---
def report_bad_topics_grouped(wb) -> str:
    ws = wb.worksheets[0]

    topic_col_idx = -1
    subj_col_idx = -1
    header_row = -1

    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
        for c_idx, val in enumerate(row):
            if isinstance(val, str):
                if "Тема урока" in val: topic_col_idx = c_idx
                if "Предмет" in val: subj_col_idx = c_idx
        if topic_col_idx != -1:
            header_row = r_idx
            break

    if topic_col_idx == -1: topic_col_idx = 5
    if subj_col_idx == -1: subj_col_idx = 2

    start_row = header_row + 2 if header_row != -1 else 2
    errors = defaultdict(list)
    count = 0

    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if len(row) <= max(topic_col_idx, subj_col_idx): continue

        subj = row[subj_col_idx]
        topic = row[topic_col_idx]
        if not subj:
            subj = "Без предмета"
        else:
            subj = str(subj).strip()

        t_str = str(topic).strip() if topic else ""
        is_bad = False

        if not topic:
            t_str = "(пустая ячейка)"
            is_bad = True
        elif not THEME_REGEX.match(t_str):
            is_bad = True

        if is_bad:
            errors[subj].append(t_str)
            count += 1

    if count == 0: return "✅ Все темы верные!"

    lines = [f"⚠️ <b>Темы с ошибками ({count} шт):</b>\n"]
    for subj in sorted(errors.keys()):
        lines.append(f"📕 <b>{subj}</b>")
        for bad_t in errors[subj]:
            lines.append(f"  • {bad_t}")
        lines.append("")
    return "\n".join(lines)


# --- Метод 3: Отчет по студентам ---
def report_students_bad_grades(wb) -> str:
    ws = wb.worksheets[0]

    fio_idx = -1
    hw_idx = -1
    cr_idx = -1
    header_row = -1

    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
        for c_idx, val in enumerate(row):
            if not isinstance(val, str): continue

            val_clean = val.strip().lower()

            if val_clean in ["fio", "фио"]:
                fio_idx = c_idx

            if val_clean in ["homework", "дз", "домашняяработа"]:
                hw_idx = c_idx

            if val_clean in ["classroom", "кр", "класснаяработа"]:
                cr_idx = c_idx

        if fio_idx != -1 and hw_idx != -1:
            header_row = r_idx
            break

    if fio_idx == -1 or hw_idx == -1 or cr_idx == -1:
        return f"❌ Не нашел нужные колонки (FIO, Homework, Classroom). Проверь заголовки."

    hw_bad_list = []
    cr_bad_list = []

    start_row = header_row + 2

    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if len(row) <= max(fio_idx, hw_idx, cr_idx): continue

        fio = row[fio_idx]
        hw_val = row[hw_idx]
        cr_val = row[cr_idx]

        if not fio: continue

        try:
            hw_score = float(hw_val)
            if hw_score <= 1.05:
                hw_bad_list.append(f"{fio} (ДЗ: {hw_val})")
        except (ValueError, TypeError):
            pass

        try:
            cr_score = float(cr_val)
            if cr_score < 3:
                cr_bad_list.append(f"{fio} (КР: {cr_val})")
        except (ValueError, TypeError):
            pass

    # если вообще всё идеально
    if not hw_bad_list and not cr_bad_list:
        return "🎉 <b>Идеально!</b> Нет студентов с ДЗ=1 или КР<3."

    report = []

    # блок ДЗ
    if hw_bad_list:
        report.append(f"📉 <b>ДЗ = 1 ({len(hw_bad_list)} чел):</b>")
        for s in hw_bad_list:
            report.append(f"  • {s}")
    else:
        report.append("✅ <b>По ДЗ (оценка 1):</b> никого не найдено.")

    report.append("")

    # блок КР
    if cr_bad_list:
        report.append(f"🆘 <b>КР меньше 3 ({len(cr_bad_list)} чел):</b>")
        for s in cr_bad_list:
            report.append(f"  • {s}")
    else:
        report.append("✅ <b>По КР (оценка меньше 3):</b> никого не найдено.")

    return "\n".join(report)


# --- Метод 4: Посещаемость по преподавателям (< 40%) ---
def report_teachers_attendance_below_40(wb, threshold=40.0) -> str:
    ws = wb.worksheets[0]

    fio_idx = -1
    avg_idx = -1
    header_row = -1

    # Ищем строку заголовков
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        for c_idx, val in enumerate(row):
            if not isinstance(val, str):
                continue
            v = val.strip().lower()

            if "фио преподавателя" in v:
                fio_idx = c_idx
            if "средняя посещаемость" in v:
                avg_idx = c_idx

        if fio_idx != -1 and avg_idx != -1:
            header_row = r_idx
            break

    if header_row == -1:
        return "❌ Не нашёл заголовки 'ФИО преподавателя' и/или 'Средняя посещаемость'."

    def to_percent(x):
        if x is None:
            return None

        # Если Excel отдал число
        if isinstance(x, (int, float)):
            val = float(x)
            return val * 100 if 0 <= val <= 1 else val

        # Если Excel отдал строку типа "73%"
        s = str(x).strip().replace("%", "").replace(",", ".")
        if not s:
            return None
        try:
            val = float(s)
            return val * 100 if 0 <= val <= 1 else val
        except ValueError:
            return None

    bad = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if len(row) <= max(fio_idx, avg_idx):
            continue

        fio = row[fio_idx]
        avg = to_percent(row[avg_idx])

        if not fio or avg is None:
            continue

        if avg <= threshold:
            bad.append((avg, str(fio).strip()))

    if not bad:
        return f"✅ <b>Посещаемость ниже {int(threshold)}%</b>: преподавателей не найдено."

    bad.sort(key=lambda x: x[0])

    lines = [f"⚠️ <b>Посещаемость ниже {int(threshold)}%:</b>\n"]
    for avg, fio in bad:
        lines.append(f"• <b>{fio}</b>: {avg:.0f}%")

    return "\n".join(lines)


# --- Главный маршрутизатор ---
def process_excel_file(data: bytes) -> str:
    if detect_excel_type(data) != "xlsx":
        return "❌ Нужен файл .xlsx"

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.worksheets[0]

        is_teachers_attendance = False
        is_students = False
        is_topics = False

        for row in ws.iter_rows(max_row=10, values_only=True):
            row_str = [str(c).strip().lower() for c in row if c]

            # Сначала проверяем посещаемость по преподавателям
            if any("фио преподавателя" in s for s in row_str) and any("средняя посещаемость" in s for s in row_str):
                is_teachers_attendance = True
                break

            # Потом — отчет по студентам
            if any("fio" in s or "фио" in s for s in row_str) and any("homework" in s for s in row_str):
                is_students = True
                break

            # Потом — темы уроков
            if any("тема урока" in s for s in row_str):
                is_topics = True
                break

        if is_teachers_attendance:
            return report_teachers_attendance_below_40(wb, threshold=40.0)
        elif is_students:
            return report_students_bad_grades(wb)
        elif is_topics:
            return report_bad_topics_grouped(wb)
        else:
            return report_schedule_count(wb)  # по дефолту пробуем расписание

    except Exception as e:
        return f"❌ Ошибка обработки: {e}"
