import io
import re
from collections import defaultdict, Counter

from telegram import Update
from telegram.constants import ParseMode
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters

import openpyxl

BOT_TOKEN = "PASTE_YOUR_TOKEN_HERE"

THEME_REGEX = re.compile(r"^Урок\s*№\s*\d+\.\s*Тема:\s*.+$", re.IGNORECASE)


def detect_excel_type(data: bytes) -> str:
    if len(data) >= 2 and data[0:2] == b"PK": return "xlsx"
    return "unknown"


# --- Метод 1: Расписание ---
def report_schedule_count(wb) -> str:
    counter = Counter()
    ws = wb.worksheets[0]
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str) and "Предмет:" in cell:
                for line in cell.splitlines():
                    line = line.strip()
                    if line.startswith("Предмет:"):
                        subj = line.replace("Предмет:", "", 1).strip()
                        if subj: counter[subj] += 1

    if not counter: return "Не нашел строк 'Предмет:'."

    lines = ["📊 <b>Количество пар по предметам:</b>\n"]
    for name, cnt in counter.most_common():
        lines.append(f"▫️ {name}: <b>{cnt}</b>")
    return "\n".join(lines)


# --- Метод 2: Темы уроков ---
def report_bad_topics_grouped(wb) -> str:
    ws = wb.worksheets[0]

    topic_col_idx = -1
    subj_col_idx = -1
    header_row = -1

    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
        for c_idx, val in enumerate(row):
            if isinstance(val, str):
                if "Тема урока" in val: topic_col_idx = c_idx
                if "Предмет" in val: subj_col_idx = c_idx
        if topic_col_idx != -1:
            header_row = r_idx
            break

    # если заголовки не нашлись
    if topic_col_idx == -1: topic_col_idx = 5  # F
    if subj_col_idx == -1: subj_col_idx = 2  # C

    start_row = header_row + 2 if header_row != -1 else 2

    errors = defaultdict(list)
    count = 0

    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if len(row) <= max(topic_col_idx, subj_col_idx): continue

        subj = row[subj_col_idx]
        topic = row[topic_col_idx]

        if not subj:
            subj = "Без предмета"
        else:
            subj = str(subj).strip()

        t_str = str(topic).strip() if topic else ""
        is_bad = False

        if not topic:
            t_str = "(пустая ячейка)"
            is_bad = True
        elif not THEME_REGEX.match(t_str):
            is_bad = True

        if is_bad:
            errors[subj].append(t_str)
            count += 1

    if count == 0: return "✅ Все темы верные!"

    lines = [f"⚠️ <b>Найдено ошибок: {count}</b>\nFormat: <i>Урок № X. Тема: Y</i>\n"]

    for subj in sorted(errors.keys()):
        lines.append(f"📕 <b>{subj}</b>")
        for bad_t in errors[subj]:
            lines.append(f"  • {bad_t}")
        lines.append("")

    return "\n".join(lines)


# --- вспомогательная функция отправки длинных сообщений ---
async def send_long_message(update: Update, text: str):
    LIMIT = 4000

    # если текст короткий - отправляем сразу
    if len(text) <= LIMIT:
        await update.message.reply_text(text, parse_mode=ParseMode.HTML)
        return

    # если длинный - режем на куски
    buffer = ""
    for line in text.splitlines(keepends=True):
        if len(buffer) + len(line) > LIMIT:
            # отправляем накопившийся кусок
            await update.message.reply_text(buffer, parse_mode=ParseMode.HTML)
            buffer = ""
        buffer += line

    # Отправляем остаток
    if buffer:
        await update.message.reply_text(buffer, parse_mode=ParseMode.HTML)


# --- Обработчик ---
async def on_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    doc = update.message.document
    await update.message.reply_text("📥 Анализирую...")

    try:
        f = await doc.get_file()
        data = await f.download_as_bytearray()

        if detect_excel_type(data) != "xlsx":
            await update.message.reply_text("❌ Нужен .xlsx файл")
            return

        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

        # авто-определение типа метода
        ws = wb.worksheets[0]
        is_topics = False
        for row in ws.iter_rows(max_row=5, values_only=True):
            for c in row:
                if isinstance(c, str) and "Тема урока" in c:
                    is_topics = True
                    break

        if is_topics:
            full_report = report_bad_topics_grouped(wb)
        else:
            full_report = report_schedule_count(wb)

        # отправляем (с разбивкой, если надо)
        await send_long_message(update, full_report)

    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка: {e}")


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("👋 Пришли .xlsx файл")


def main():
    app = ApplicationBuilder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.Document.ALL, on_document))
    print("Бот работает...")
    app.run_polling()


if __name__ == "__main__":
    main()
